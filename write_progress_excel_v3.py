import json
import openpyxl
import datetime
import re
import os
import sys

def strip_emojis(text):
    if not text:
        return ""
    # Regex to match emojis, dingbats, and other supplemental symbols
    emoji_pattern = re.compile(
        r'[\U00010000-\U0010FFFF\u2600-\u27BF\u2300-\u23FF\u2B50\u2B06\u2192\u26A0\u26A1]',
        re.UNICODE
    )
    cleaned = emoji_pattern.sub('', text)
    return cleaned.strip()

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    
    excel_path = r"c:\Users\admin\Desktop\Threads\水果市場評論貼文_匯入進度.xlsx"
    json_path = r"c:\Users\admin\Desktop\Threads\extracted_posts_all_v2.json"
    
    if not os.path.exists(json_path):
        print(f"Error: JSON file not found at {json_path}")
        sys.exit(1)
        
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)
        
    print(f"Reading V2 JSON data from {json_path}...")
    with open(json_path, 'r', encoding='utf-8') as f:
        posts = json.load(f)
        
    print(f"Opening Excel file: {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    print(f"Active sheet: {sheet.title}, Current rows: {sheet.max_row}")
    
    # Read existing full contents to prevent duplicates
    existing_full_contents = set()
    for row in range(2, sheet.max_row + 1):
        cell_val = sheet.cell(row=row, column=6).value # Col 6 is "貼文內容全文"
        if cell_val:
            existing_full_contents.add(cell_val.strip())
            
    print(f"Loaded {len(existing_full_contents)} existing posts to check for duplicates.")
    
    appended_count = 0
    for idx, post in enumerate(posts):
        full_content = post['fullContent'].strip()
        
        # Check duplicate by matching first 100 characters of fullContent
        is_duplicate = False
        for ext in existing_full_contents:
            if full_content[:100] in ext or ext[:100] in full_content:
                is_duplicate = True
                break
                
        if is_duplicate:
            print(f"  -> Skipping duplicate post: {post['smallTitle'][:30]}...")
            continue
            
        # Parse timestamp and clear hour/minute/second to match formatting
        dt_val = datetime.datetime.fromtimestamp(post['dateTimestamp'])
        dt_val_date_only = dt_val.replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Clean smallTitle by removing emojis
        cleaned_small_title = strip_emojis(post['smallTitle'])
        
        row_data = [
            post['username'],         # Col 1: 作者
            post['authorName'],       # Col 2: 貼文大標題
            cleaned_small_title,      # Col 3: 貼文小標題 (emoji-free)
            dt_val_date_only,         # Col 4: 發布日期 (date only)
            post['originalContent'],  # Col 5: 貼文內容
            post['fullContent']       # Col 6: 貼文內容全文
        ]
        
        sheet.append(row_data)
        appended_count += 1
        print(f"  -> Appended: {cleaned_small_title} by {post['authorName']}")
        
    if appended_count > 0:
        print(f"Saving Excel file: {excel_path}...")
        wb.save(excel_path)
        print("Excel file updated and saved successfully!")
    else:
        print("No new posts to append. Excel file was not modified.")

if __name__ == "__main__":
    main()
