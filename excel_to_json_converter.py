import json
import openpyxl
import os
import sys

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    current_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(current_dir, "水果市場評論貼文_匯入進度.xlsx")
    json_output_path = os.path.join(current_dir, "水果市場評論貼文_匯入進度.json")
    html_output_path = os.path.join(current_dir, "水果市場評論貼文_瀏覽.html")
    index_output_path = os.path.join(current_dir, "index.html")
    
    if not os.path.exists(excel_path):
        print(f"錯誤：找不到 Excel 檔案 {excel_path}")
        return

    print("正在讀取 Excel 檔案...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        
        posts = []
        # Columns: 作者 (1), 貼文大標題 (2), 貼文小標題 (3), 發布日期 (4), 貼文內容 (5), 貼文內容全文 (6)
        for row in range(2, sheet.max_row + 1):
            username = sheet.cell(row=row, column=1).value
            authorName = sheet.cell(row=row, column=2).value
            smallTitle = sheet.cell(row=row, column=3).value
            date_val = sheet.cell(row=row, column=4).value
            originalContent = sheet.cell(row=row, column=5).value
            fullContent = sheet.cell(row=row, column=6).value
            
            # Skip empty rows
            if not username and not smallTitle and not fullContent:
                continue
                
            # Format date
            date_str = ""
            if date_val:
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val).split(' ')[0]
            
            posts.append({
                "username": str(username or "").strip(),
                "authorName": str(authorName or "").strip(),
                "smallTitle": str(smallTitle or "").strip(),
                "date": date_str,
                "originalContent": str(originalContent or "").strip(),
                "fullContent": str(fullContent or "").strip()
            })
            
        # Sort by date desc
        try:
            posts.sort(key=lambda x: x['date'], reverse=True)
        except Exception:
            pass

        # Write to JSON file
        print(f"正在寫入 JSON 檔案：{json_output_path}...")
        with open(json_output_path, 'w', encoding='utf-8') as f:
            json.dump(posts, f, ensure_ascii=False, indent=2)
            
        # Write to HTML file
        print(f"正在生成瀏覽網頁：{html_output_path}...")
        html_content = generate_html(posts)
        with open(html_output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        # Write to index.html for GitHub Pages hosting
        print(f"正在生成 GitHub Pages 首頁：{index_output_path}...")
        with open(index_output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        print("轉換成功！您現在可以直接雙擊開啟 '水果市場評論貼文_瀏覽.html' 進行閱讀。")
        
    except Exception as e:
        print(f"發生錯誤：{str(e)}")

def generate_html(posts_data):
    # JSON dump for embedding
    embedded_json = json.dumps(posts_data, ensure_ascii=False, indent=2)
    
    html_template = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>水果市場評論貼文 - 數據瀏覽器</title>
    <style>
        :root {{
            --bg-color: #f8fafc;
            --card-bg: #ffffff;
            --border-color: #e2e8f0;
            --accent: #4f46e5;
            --accent-glow: rgba(79, 70, 229, 0.1);
            --text-primary: #0f172a;
            --text-secondary: #475569;
            --text-muted: #64748b;
        }}
        
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        
        body {{
            background-color: var(--bg-color);
            color: var(--text-primary);
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            line-height: 1.6;
            padding: 40px 20px;
        }}
        
        .container {{
            max-width: 1100px;
            margin: 0 auto;
            display: flex;
            flex-direction: column;
            gap: 24px;
        }}
        
        header {{
            text-align: center;
            margin-bottom: 10px;
        }}
        
        h1 {{
            font-size: 2rem;
            font-weight: 700;
            color: #0f172a;
            margin-bottom: 8px;
        }}
        
        .subtitle {{
            color: var(--text-secondary);
            font-size: 0.95rem;
        }}
        
        /* Controls */
        .controls-card {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            display: flex;
            gap: 16px;
            align-items: center;
            flex-wrap: wrap;
        }}
        
        .search-group {{
            flex: 1;
            min-width: 250px;
        }}
        
        .search-input {{
            width: 100%;
            background: #ffffff;
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            padding: 10px 16px;
            color: var(--text-primary);
            font-size: 0.9rem;
            outline: none;
        }}
        
        .search-input:focus {{
            border-color: var(--accent);
            box-shadow: 0 0 0 3px var(--accent-glow);
        }}
        
        .sort-group select {{
            background: #ffffff;
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            padding: 10px 16px;
            color: var(--text-primary);
            font-size: 0.9rem;
            outline: none;
            cursor: pointer;
        }}
        
        .stats-badge {{
            background: #f1f5f9;
            border: 1px solid #e2e8f0;
            color: var(--text-secondary);
            padding: 8px 16px;
            border-radius: 8px;
            font-size: 0.9rem;
            font-weight: 600;
        }}
        
        /* Table layout */
        .table-container {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        
        th, td {{
            padding: 16px;
            text-align: left;
            border-bottom: 1px solid var(--border-color);
            font-size: 0.95rem;
        }}
        
        th {{
            background-color: #f8fafc;
            font-weight: 600;
            color: var(--text-primary);
            border-bottom: 2px solid var(--border-color);
        }}
        
        tr.data-row:hover td {{
            background-color: #f8fafc;
        }}
        
        .author-badge {{
            font-weight: 500;
            color: var(--accent);
        }}
        
        .post-title {{
            font-weight: 600;
            color: var(--text-primary);
        }}
        
        .post-content {{
            color: var(--text-secondary);
            white-space: pre-wrap;
            word-break: break-word;
            font-size: 0.9rem;
        }}
        
        .expand-btn {{
            background: #f1f5f9;
            border: 1px solid #e2e8f0;
            color: var(--text-secondary);
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 0.8rem;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            gap: 4px;
            margin-top: 6px;
            outline: none;
        }}
        
        .expand-btn:hover {{
            background: #e2e8f0;
            color: var(--text-primary);
        }}
        
        .full-content-box {{
            margin-top: 10px;
            background: #f8fafc;
            border: 1px dashed #cbd5e1;
            border-radius: 8px;
            padding: 12px;
            font-size: 0.85rem;
            color: var(--text-primary);
            white-space: pre-wrap;
        }}
        
        .hidden {{
            display: none !important;
        }}
        
        .empty-feed {{
            text-align: center;
            padding: 40px;
            color: var(--text-muted);
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>水果市場評論貼文 - 數據瀏覽器</h1>
            <p class="subtitle">讀取自本地 Excel 檔案轉換之 JSON 數據</p>
        </header>
        
        <div class="controls-card">
            <div class="search-group">
                <input type="text" id="search-input" class="search-input" placeholder="搜尋日期、標題、作者或內容...">
            </div>
            
            <div class="sort-group">
                <select id="sort-select">
                    <option value="newest">發布日期: 由新到舊</option>
                    <option value="oldest">發布日期: 由舊到新</option>
                    <option value="title">標題排序</option>
                </select>
            </div>
            
            <div class="stats-badge">
                共有 <span id="posts-count">0</span> 筆紀錄
            </div>
        </div>
        
        <div class="table-container">
            <table id="posts-table">
                <thead>
                    <tr>
                        <th style="width: 120px;">發布日期</th>
                        <th style="width: 150px;">作者</th>
                        <th style="width: 220px;">貼文標題</th>
                        <th>貼文內容</th>
                    </tr>
                </thead>
                <tbody id="table-body"></tbody>
            </table>
            <div id="empty-state" class="empty-feed hidden">
                沒有找到符合條件的貼文紀錄。
            </div>
        </div>
    </div>

    <script>
        const postsData = {embedded_json};

        const searchInput = document.getElementById('search-input');
        const sortSelect = document.getElementById('sort-select');
        const postsCount = document.getElementById('posts-count');
        const tableBody = document.getElementById('table-body');
        const emptyState = document.getElementById('empty-state');

        // Initial load
        applyFiltersAndSort();

        // Listeners
        searchInput.addEventListener('input', applyFiltersAndSort);
        sortSelect.addEventListener('change', applyFiltersAndSort);

        function applyFiltersAndSort() {{
            const searchText = searchInput.value.toLowerCase().trim();
            const sortBy = sortSelect.value;
            
            // 1. Filter
            let filtered = postsData.filter(post => {{
                return (
                    post.smallTitle.toLowerCase().includes(searchText) ||
                    post.originalContent.toLowerCase().includes(searchText) ||
                    post.fullContent.toLowerCase().includes(searchText) ||
                    post.authorName.toLowerCase().includes(searchText) ||
                    post.username.toLowerCase().includes(searchText) ||
                    post.date.includes(searchText)
                );
            }});
            
            // 2. Sort
            filtered.sort((a, b) => {{
                if (sortBy === 'newest') {{
                    return new Date(b.date) - new Date(a.date);
                }} else if (sortBy === 'oldest') {{
                    return new Date(a.date) - new Date(b.date);
                }} else if (sortBy === 'title') {{
                    return a.smallTitle.localeCompare(b.smallTitle, 'zh-TW');
                }}
                return 0;
            }});
            
            postsCount.innerText = filtered.length;
            renderTable(filtered);
        }}

        function renderTable(posts) {{
            tableBody.innerHTML = '';
            
            if (posts.length === 0) {{
                emptyState.classList.remove('hidden');
                document.getElementById('posts-table').classList.add('hidden');
                return;
            }}
            
            emptyState.classList.add('hidden');
            document.getElementById('posts-table').classList.remove('hidden');
            
            posts.forEach((post, idx) => {{
                const tr = document.createElement('tr');
                tr.className = 'data-row';
                
                const hasFullContent = post.fullContent && post.fullContent.trim() !== post.originalContent.trim();
                
                tr.innerHTML = `
                    <td><strong>${{escapeHtml(post.date)}}</strong></td>
                    <td>
                        <span class="author-badge">@${{escapeHtml(post.username)}}</span>
                        <div style="font-size: 0.8rem; color: var(--text-muted);">${{escapeHtml(post.authorName)}}</div>
                    </td>
                    <td><div class="post-title">${{escapeHtml(post.smallTitle)}}</div></td>
                    <td>
                        <div class="post-content">${{formatText(post.originalContent)}}</div>
                        ${{hasFullContent ? `
                            <button class="expand-btn" onclick="toggleFullContent(this, ${{idx}})">
                                <span>展開完整子貼文串 ▼</span>
                            </button>
                            <div id="full-box-${{idx}}" class="full-content-box hidden">${{formatText(post.fullContent)}}</div>
                        ` : ''}}
                    </td>
                `;
                tableBody.appendChild(tr);
            }});
        }}

        function toggleFullContent(btn, idx) {{
            const box = document.getElementById(`full-box-${{idx}}`);
            if (box.classList.contains('hidden')) {{
                box.classList.remove('hidden');
                btn.querySelector('span').innerText = '收起完整子貼文串 ▲';
            }} else {{
                box.classList.add('hidden');
                btn.querySelector('span').innerText = '展開完整子貼文串 ▼';
            }}
        }}

        function formatText(text) {{
            return escapeHtml(text).replace(/\\n/g, '<br>');
        }}

        function escapeHtml(unsafe) {{
            return unsafe
                 .replace(/&/g, "&amp;")
                 .replace(/</g, "&lt;")
                 .replace(/>/g, "&gt;")
                 .replace(/"/g, "&quot;")
                 .replace(/'/g, "&#039;");
        }}
    </script>
</body>
</html>
"""
    return html_template

if __name__ == "__main__":
    main()
